#!/usr/bin/python3

import numpy as np
from array import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
import csv
import sys
import subprocess
import os
import shlex
import datetime, time
import mysql.connector as mariadb
import unicodedata
import codecs
import io
import string

#reload(sys)
#sys.setdefaultencoding("utf-8")

### Open Excel Worksheet with default styles ###
wb=Workbook()
ws=wb.active

alignment=Alignment(horizontal='general',
		      vertical='bottom',
		      text_rotation=0,
		      wrap_text=False,
		      shrink_to_fit=False,
		      indent=0)

font = Font(name='Calibri',
                 size=10,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')
###   ###

inpfileb=sys.argv[1]
outfileb=sys.argv[2]
outfile=sys.argv[3]
chid=sys.argv[4]

count=[]
progname=[]
year=[]
month=[]
day=[]
hour=[]
minute=[]
dur=[]
isot=[]
uepoch=[]
epgid=[]
progid=[]
row=[]
progtitle=[]
cupcnt=0
fname=0
fdate=0

### Take the date from input file name ###
ifnsplit=inpfileb.split('-',1)
fname=ifnsplit[0]
fdate=ifnsplit[1]
#print(fname)
#print(fdate)
###   ###


#bcmd='cat ' + inpfileb + ' |grep ' + chid+'_ |awk \'{print $13}\' |cut -d \'/\' -f 4 |cut -b 5- |sort |uniq -c > ' + outfileb + '\n'
bcmd='cat ' + inpfileb + ' |grep -v \'STATUS: 404\' ' + ' |grep ' + chid+'_ |awk \'{print $13}\' |cut -d \'/\' -f 4 |cut -b 5- |sort |uniq -c > ' + outfileb + '\n'
#args=shlex.split(bcmd)

#print (args)
#print(bcmd)

bp=subprocess.Popen(bcmd, shell=True, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
output, error=bp.communicate()

print ('bash exit code is', bp.returncode)
#print output


with open(outfileb) as inpf:
	for line in inpf:
		split = line.strip().split(' ',1)
		count.append(split[0])
		progname.append(split[1])

cupcnt=len(progname)
print (cupcnt)

for i in range(0, len(progname)):
	splitdt = progname[i].split('_',5)
	year.append(splitdt[0])
	month.append(splitdt[1])
	day.append(splitdt[2])
	hour.append(splitdt[3])
	minute.append(splitdt[4])
	dur.append(splitdt[5])
	isot.append(datetime.datetime(int(year[i]), int(month[i]), int(day[i]), int(hour[i]), int(minute[i])))
	uepoch.append(time.mktime(isot[i].timetuple())*1000)
	epgid.append(chid+str(int(uepoch[i])))


print (year[-1], month[-1], day[-1], hour[-1], minute[-1], dur[-1])
print (isot[-1], int(uepoch[-1]))
print (progname[-1], count[-1])
print (epgid[-1])

mariadb_connection = mariadb.connect(host='example_host_ip', user='example_user', password='example_password', database='example_db', charset='utf8', collation='utf8_general_ci', autocommit='False')
cursor=mariadb_connection.cursor()
for i in range(0,len(epgid)):
   try:
      cursor.execute("select title from db.programme_title where locale='en_US' and programme_id=%s", (epgid[i],))
      row=cursor.fetchone()
      if row is None:
         mydata='Missing'
      else:
         mydata=row[0]

#      print ((u'row: %s') % mydata)
      progtitle.append(str(mydata))
#
#      progtitle.append(str(row).strip("\'\t\,\(\)\\t\\xa0"))
#
###      print((u'progtitle: %s') % progtitle[i])
   except mariadb.Error as error:
      print("Error: {}".format(error))

cursor.close()
mariadb_connection.commit()
mariadb_connection.close()

#print((u'progtitle: %s') % progtitle)

#print inpfileb
#print outfileb
print (len(epgid))
print (len(progtitle))

#uprogtitle=[unicode(item) for item in progtitle]

### Old tab delimited text file output ###
#with io.open(outfile, 'w', encoding='utf-8') as outf:
##	writer=csv.writer(outf, delimiter='\t')
#	for j in range(0,len(progtitle)):
## 		for urow in zip(progtitle, isot, count):
##		unicode(urow,'utf_8','replace')
#		outf.write("'{}'\t {}\t {}\n".format(progtitle[j], isot[j], count[j]))
#
#outf.close()
###   ###

### Excel sheet name and column titles and information styles ###
ws.title=chid+'_'+'epg_date_count'+'_'+fdate
ws['A1']="Name of the CUP Program"
ws['B1']="Date of the CUP Program"
ws['C1']="CUP Count"
ws['D1']="CUP Activity Date: " + fdate + "; CHID: " + chid
a1=ws['A1']
b1=ws['B1']
c1=ws['C1']
d1=ws['D1']
a1.font=Font(name='Calibri',size=11,bold=True)
a1.alignment=alignment
b1.font=Font(name='Calibri',size=11,bold=True)
b1.alignment=alignment
c1.font=Font(name='Calibri',size=11,bold=True)
c1.alignment=alignment
d1.font=Font(name='Calibri',size=11,bold=True,italic=True)
d1.alignment=alignment

#col = ws.column_dimensions['A']
#col.font=font
#row = ws.row_dimensions[1]
#row.font=font

###   ###

### Excel Data ###
for j in range(0,len(progtitle)):
	ws.cell(column=1, row=j+2, value=format(progtitle[j]))
	ws.cell(column=2, row=j+2, value=format(isot[j]))
	ws.cell(column=3, row=j+2, value=format(count[j]))
###   ###

### Column widths ###
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column, 0), len(cell.value)))
for col, value in dims.items():
    ws.column_dimensions[col].width = value
###   ###

wb.save(filename=outfile)
